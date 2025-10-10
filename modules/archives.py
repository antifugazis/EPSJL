from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, abort, make_response
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, ArchiveDossier, ArchiveFichier
from datetime import datetime, timedelta
import os
import uuid
from functools import wraps
import mimetypes

# Export libraries
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
# PDF and DOCX exports disabled due to Linux compatibility issues
# from reportlab.lib.pagesizes import letter, A4
# from reportlab.lib import colors
# from reportlab.lib.units import inch
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
# from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
# from docx import Document as DocxDocument
# from docx.shared import Inches, Pt, RGBColor

archives_blueprint = Blueprint('archives', __name__, url_prefix='/archives')

# Configuration pour les uploads
UPLOAD_FOLDER = 'uploads/archives'
ALLOWED_EXTENSIONS = {
    'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx',
    'txt', 'csv', 'odt', 'ods', 'odp',
    'jpg', 'jpeg', 'png', 'gif', 'bmp', 'svg', 'webp',
    'mp3', 'mp4', 'wav', 'avi', 'mov', 'wmv', 'flv',
    'zip', 'rar', '7z', 'tar', 'gz'
}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_size_str(size_bytes):
    """Convertit la taille en bytes en format lisible"""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.2f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.2f} MB"
    else:
        return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ['admin', 'directeur']:
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

# Route principale - Liste des archives
@archives_blueprint.route('/')
@login_required
def index():
    filtre = request.args.get('filtre', 'tous')
    
    # Base query - exclure les dossiers supprimés
    query = ArchiveDossier.query.filter_by(supprime=False)
    
    # Appliquer les filtres
    if filtre == 'recent':
        # Ajoutés récemment (≤ 1 mois)
        one_month_ago = datetime.now() - timedelta(days=30)
        query = query.filter(ArchiveDossier.date_creation >= one_month_ago)
    elif filtre == 'modifies':
        # Modifiés récemment
        one_month_ago = datetime.now() - timedelta(days=30)
        query = query.filter(ArchiveDossier.date_modification >= one_month_ago)
    elif filtre == 'confidentiel':
        query = query.filter_by(confidentiel=True)
    
    dossiers = query.order_by(ArchiveDossier.date_creation.desc()).all()
    
    # Compter les éléments pour chaque filtre
    counts = {
        'tous': ArchiveDossier.query.filter_by(supprime=False).count(),
        'recent': ArchiveDossier.query.filter_by(supprime=False).filter(
            ArchiveDossier.date_creation >= datetime.now() - timedelta(days=30)
        ).count(),
        'modifies': ArchiveDossier.query.filter_by(supprime=False).filter(
            ArchiveDossier.date_modification >= datetime.now() - timedelta(days=30)
        ).count(),
        'confidentiel': ArchiveDossier.query.filter_by(supprime=False, confidentiel=True).count()
    }
    
    return render_template('archives/index.html', 
                         dossiers=dossiers, 
                         filtre_actuel=filtre,
                         counts=counts)

# Route pour ajouter un dossier
@archives_blueprint.route('/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter():
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            nom = request.form.get('nom')
            nombre_fichiers = int(request.form.get('nombre_fichiers', 0))
            informations_supplementaires = request.form.get('informations_supplementaires')
            sauvegarde_serveur = request.form.get('sauvegarde_serveur') == 'true'
            confidentiel = request.form.get('confidentiel') == 'true'
            code_pin = request.form.get('code_pin')
            
            # Validation
            if not nom:
                flash('Le nom du dossier est obligatoire', 'error')
                return redirect(url_for('archives.ajouter'))
            
            # Gérer la photo de couverture du dossier
            photo_couverture = None
            if 'photo_couverture' in request.files:
                file = request.files['photo_couverture']
                if file and file.filename:
                    if allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        unique_filename = f"{uuid.uuid4().hex}_{filename}"
                        
                        # Créer le dossier s'il n'existe pas
                        os.makedirs(os.path.join(UPLOAD_FOLDER, 'couvertures'), exist_ok=True)
                        
                        file_path = os.path.join(UPLOAD_FOLDER, 'couvertures', unique_filename)
                        file.save(file_path)
                        photo_couverture = file_path
            
            # Créer le dossier
            nouveau_dossier = ArchiveDossier(
                nom=nom,
                photo_couverture=photo_couverture,
                nombre_fichiers=nombre_fichiers,
                informations_supplementaires=informations_supplementaires,
                sauvegarde_serveur=sauvegarde_serveur,
                confidentiel=confidentiel,
                code_pin=generate_password_hash(code_pin) if confidentiel and code_pin else None,
                cree_par=current_user.id
            )
            
            db.session.add(nouveau_dossier)
            db.session.flush()  # Pour obtenir l'ID du dossier
            
            # Ajouter les fichiers
            for i in range(nombre_fichiers):
                nom_document = request.form.get(f'fichier_{i}_nom')
                note_additionnelle = request.form.get(f'fichier_{i}_note')
                
                if not nom_document:
                    continue
                
                # Photo de couverture du fichier
                photo_fichier = None
                if f'fichier_{i}_photo' in request.files:
                    file = request.files[f'fichier_{i}_photo']
                    if file and file.filename:
                        if allowed_file(file.filename):
                            filename = secure_filename(file.filename)
                            unique_filename = f"{uuid.uuid4().hex}_{filename}"
                            
                            os.makedirs(os.path.join(UPLOAD_FOLDER, 'fichiers_couvertures'), exist_ok=True)
                            
                            file_path = os.path.join(UPLOAD_FOLDER, 'fichiers_couvertures', unique_filename)
                            file.save(file_path)
                            photo_fichier = file_path
                
                # Fichier principal
                if f'fichier_{i}_document' in request.files:
                    file = request.files[f'fichier_{i}_document']
                    if file and file.filename:
                        if allowed_file(file.filename):
                            filename = secure_filename(file.filename)
                            unique_filename = f"{uuid.uuid4().hex}_{filename}"
                            
                            # Créer le dossier pour ce dossier d'archives
                            dossier_path = os.path.join(UPLOAD_FOLDER, 'documents', str(nouveau_dossier.id))
                            os.makedirs(dossier_path, exist_ok=True)
                            
                            file_path = os.path.join(dossier_path, unique_filename)
                            file.save(file_path)
                            
                            # Obtenir la taille et le type du fichier
                            file_size = os.path.getsize(file_path)
                            file_type = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'unknown'
                            
                            # Créer l'entrée de fichier
                            nouveau_fichier = ArchiveFichier(
                                dossier_id=nouveau_dossier.id,
                                nom_document=nom_document,
                                photo_couverture=photo_fichier,
                                fichier_path=file_path,
                                fichier_type=file_type,
                                fichier_taille=file_size,
                                note_additionnelle=note_additionnelle
                            )
                            
                            db.session.add(nouveau_fichier)
            
            db.session.commit()
            flash(f'Le dossier "{nom}" a été créé avec succès', 'success')
            return redirect(url_for('archives.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la création du dossier: {str(e)}', 'error')
            return redirect(url_for('archives.ajouter'))
    
    return render_template('archives/ajouter.html')

# Route pour voir les détails d'un dossier
@archives_blueprint.route('/details/<int:dossier_id>')
@login_required
def details(dossier_id):
    dossier = ArchiveDossier.query.get_or_404(dossier_id)
    
    # Vérifier si le dossier est confidentiel
    if dossier.confidentiel:
        # Vérifier si l'utilisateur a déjà déverrouillé ce dossier dans cette session
        session_key = f'archive_unlocked_{dossier_id}'
        if not request.args.get('unlocked') and session_key not in request.cookies:
            return render_template('archives/verifier_pin.html', dossier=dossier)
    
    fichiers = ArchiveFichier.query.filter_by(dossier_id=dossier_id).order_by(ArchiveFichier.date_ajout.desc()).all()
    
    # Ajouter la taille formatée pour chaque fichier
    for fichier in fichiers:
        fichier.taille_formatee = get_file_size_str(fichier.fichier_taille) if fichier.fichier_taille else 'N/A'
    
    response = render_template('archives/details.html', dossier=dossier, fichiers=fichiers)
    
    # Si le dossier vient d'être déverrouillé, définir un cookie
    if request.args.get('unlocked'):
        response = make_response(response)
        response.set_cookie(f'archive_unlocked_{dossier_id}', 'true', max_age=3600)  # 1 heure
    
    return response

# Route pour vérifier le code PIN
@archives_blueprint.route('/verifier-pin/<int:dossier_id>', methods=['POST'])
@login_required
def verifier_pin(dossier_id):
    dossier = ArchiveDossier.query.get_or_404(dossier_id)
    
    if not dossier.confidentiel:
        return redirect(url_for('archives.details', dossier_id=dossier_id))
    
    code_pin = request.form.get('code_pin')
    
    if check_password_hash(dossier.code_pin, code_pin):
        return redirect(url_for('archives.details', dossier_id=dossier_id, unlocked=True))
    else:
        flash('Code PIN incorrect', 'error')
        return redirect(url_for('archives.details', dossier_id=dossier_id))

# Route pour modifier un dossier
@archives_blueprint.route('/modifier/<int:dossier_id>', methods=['GET', 'POST'])
@login_required
def modifier(dossier_id):
    dossier = ArchiveDossier.query.get_or_404(dossier_id)
    
    if request.method == 'POST':
        try:
            dossier.nom = request.form.get('nom')
            dossier.informations_supplementaires = request.form.get('informations_supplementaires')
            dossier.sauvegarde_serveur = request.form.get('sauvegarde_serveur') == 'true'
            dossier.confidentiel = request.form.get('confidentiel') == 'true'
            
            # Mettre à jour le code PIN si nécessaire
            code_pin = request.form.get('code_pin')
            if dossier.confidentiel and code_pin:
                dossier.code_pin = generate_password_hash(code_pin)
            elif not dossier.confidentiel:
                dossier.code_pin = None
            
            # Gérer la nouvelle photo de couverture
            if 'photo_couverture' in request.files:
                file = request.files['photo_couverture']
                if file and file.filename:
                    if allowed_file(file.filename):
                        # Supprimer l'ancienne photo si elle existe
                        if dossier.photo_couverture and os.path.exists(dossier.photo_couverture):
                            os.remove(dossier.photo_couverture)
                        
                        filename = secure_filename(file.filename)
                        unique_filename = f"{uuid.uuid4().hex}_{filename}"
                        
                        os.makedirs(os.path.join(UPLOAD_FOLDER, 'couvertures'), exist_ok=True)
                        
                        file_path = os.path.join(UPLOAD_FOLDER, 'couvertures', unique_filename)
                        file.save(file_path)
                        dossier.photo_couverture = file_path
            
            dossier.date_modification = datetime.now()
            db.session.commit()
            
            flash(f'Le dossier "{dossier.nom}" a été modifié avec succès', 'success')
            return redirect(url_for('archives.details', dossier_id=dossier_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la modification: {str(e)}', 'error')
    
    return render_template('archives/modifier.html', dossier=dossier)

# Route pour supprimer un dossier (mise à la corbeille)
@archives_blueprint.route('/supprimer/<int:dossier_id>', methods=['POST'])
@login_required
def supprimer(dossier_id):
    dossier = ArchiveDossier.query.get_or_404(dossier_id)
    
    try:
        dossier.supprime = True
        dossier.date_suppression = datetime.now()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Le dossier "{dossier.nom}" a été déplacé vers la corbeille'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Route pour la corbeille
@archives_blueprint.route('/corbeille')
@login_required
def corbeille():
    # Dossiers supprimés depuis moins de 30 jours
    thirty_days_ago = datetime.now() - timedelta(days=30)
    dossiers = ArchiveDossier.query.filter(
        ArchiveDossier.supprime == True,
        ArchiveDossier.date_suppression >= thirty_days_ago
    ).order_by(ArchiveDossier.date_suppression.desc()).all()
    
    # Calculer les jours restants pour chaque dossier
    for dossier in dossiers:
        if dossier.date_suppression:
            jours_restants = 30 - (datetime.now() - dossier.date_suppression).days
            dossier.jours_restants = max(0, jours_restants)
    
    return render_template('archives/corbeille.html', dossiers=dossiers)

# Route pour restaurer un dossier
@archives_blueprint.route('/restaurer/<int:dossier_id>', methods=['POST'])
@login_required
def restaurer(dossier_id):
    dossier = ArchiveDossier.query.get_or_404(dossier_id)
    
    try:
        dossier.supprime = False
        dossier.date_suppression = None
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Le dossier "{dossier.nom}" a été restauré'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Route pour supprimer définitivement un dossier
@archives_blueprint.route('/supprimer-definitivement/<int:dossier_id>', methods=['POST'])
@login_required
@admin_required
def supprimer_definitivement(dossier_id):
    dossier = ArchiveDossier.query.get_or_404(dossier_id)
    
    try:
        # Supprimer les fichiers physiques
        for fichier in dossier.fichiers:
            if fichier.fichier_path and os.path.exists(fichier.fichier_path):
                os.remove(fichier.fichier_path)
            if fichier.photo_couverture and os.path.exists(fichier.photo_couverture):
                os.remove(fichier.photo_couverture)
        
        # Supprimer la photo de couverture du dossier
        if dossier.photo_couverture and os.path.exists(dossier.photo_couverture):
            os.remove(dossier.photo_couverture)
        
        # Supprimer le dossier de la base de données
        nom_dossier = dossier.nom
        db.session.delete(dossier)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Le dossier "{nom_dossier}" a été supprimé définitivement'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Route pour télécharger un fichier
@archives_blueprint.route('/telecharger/<int:fichier_id>')
@login_required
def telecharger(fichier_id):
    fichier = ArchiveFichier.query.get_or_404(fichier_id)
    
    # Vérifier si le dossier parent est confidentiel
    if fichier.dossier.confidentiel:
        session_key = f'archive_unlocked_{fichier.dossier_id}'
        if session_key not in request.cookies:
            abort(403)
    
    if not os.path.exists(fichier.fichier_path):
        abort(404)
    
    return send_file(fichier.fichier_path, as_attachment=True, download_name=fichier.nom_document)

# Routes d'export
# PDF export disabled due to Linux compatibility issues with reportlab
# @archives_blueprint.route('/export/pdf')
# @login_required
# def export_pdf():
#     flash('Export PDF temporairement désactivé. Utilisez Excel à la place.', 'warning')
#     return redirect(url_for('archives.index'))

@archives_blueprint.route('/export/excel')
@login_required
def export_excel():
    filtre = request.args.get('filtre', 'tous')
    
    # Récupérer les dossiers selon le filtre
    query = ArchiveDossier.query.filter_by(supprime=False)
    
    if filtre == 'recent':
        one_month_ago = datetime.now() - timedelta(days=30)
        query = query.filter(ArchiveDossier.date_creation >= one_month_ago)
    elif filtre == 'modifies':
        one_month_ago = datetime.now() - timedelta(days=30)
        query = query.filter(ArchiveDossier.date_modification >= one_month_ago)
    elif filtre == 'confidentiel':
        query = query.filter_by(confidentiel=True)
    
    dossiers = query.order_by(ArchiveDossier.date_creation.desc()).all()
    
    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Archives"
    
    # Style pour l'en-tête
    header_fill = PatternFill(start_color="00AEEF", end_color="00AEEF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # En-têtes
    headers = ['Nom du Dossier', 'Date de Création', 'Nombre de Fichiers', 'Confidentiel', 'Créé par', 'Informations']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Données
    for row, dossier in enumerate(dossiers, 2):
        ws.cell(row=row, column=1, value=dossier.nom)
        ws.cell(row=row, column=2, value=dossier.date_creation.strftime('%d/%m/%Y %H:%M') if dossier.date_creation else 'N/A')
        ws.cell(row=row, column=3, value=dossier.nombre_fichiers)
        ws.cell(row=row, column=4, value='Oui' if dossier.confidentiel else 'Non')
        ws.cell(row=row, column=5, value=dossier.createur.username if dossier.createur else 'N/A')
        ws.cell(row=row, column=6, value=dossier.informations_supplementaires or '')
    
    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, as_attachment=True, download_name='archives.xlsx', 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# DOCX export disabled due to Linux compatibility issues with python-docx
# @archives_blueprint.route('/export/docx')
# @login_required
# def export_docx():
#     flash('Export DOCX temporairement désactivé. Utilisez Excel à la place.', 'warning')
#     return redirect(url_for('archives.index'))

# Nettoyage automatique de la corbeille (à appeler périodiquement)
def nettoyer_corbeille():
    """Supprime définitivement les dossiers dans la corbeille depuis plus de 30 jours"""
    thirty_days_ago = datetime.now() - timedelta(days=30)
    dossiers_a_supprimer = ArchiveDossier.query.filter(
        ArchiveDossier.supprime == True,
        ArchiveDossier.date_suppression < thirty_days_ago
    ).all()
    
    for dossier in dossiers_a_supprimer:
        # Supprimer les fichiers physiques
        for fichier in dossier.fichiers:
            if fichier.fichier_path and os.path.exists(fichier.fichier_path):
                os.remove(fichier.fichier_path)
            if fichier.photo_couverture and os.path.exists(fichier.photo_couverture):
                os.remove(fichier.photo_couverture)
        
        if dossier.photo_couverture and os.path.exists(dossier.photo_couverture):
            os.remove(dossier.photo_couverture)
        
        db.session.delete(dossier)
    
    db.session.commit()
